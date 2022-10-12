from ast import Mod
from gettext import find
import requests
from bs4 import BeautifulSoup
import pandas as pd
import xlwt
import openpyxl
import random
from xlwt import Workbook

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

col = ["Marka", "Model Adı", "Model No", "İşletim Sistemi", "İşlemci Tipi", "İslemci Nesli", "Ram", "Disk Boyutu", "Disk Türü", "Ekran Boyutu", "Puanı", "Fiyat", "Site İsmi", "Site Linki"]
teknosa = "https://www.teknosa.com/laptop-notebook-c-116004?s=%3Arelevance&page={0}"
tekno = "https://www.teknosa.com"
OS = " "
cpuType = " "
cpuStatus = " "
ram = " "
Disk = " "
DiskType = " "
screen = " "
Marka = " "
pointss = ["0,0", "3,2", "3,5", "3,7", "3,9", "4,0", "4,1", "4,4", "4,5", "4,7", "5,0"]
Ozellik_adi2 = []
Ozellik_aciklamasi2 = []
Link_two = []
row = 1

for i in range(14):
    c1 = my_sheet.cell(row = 1, column = i + 1)
    c1.value = col[i]

def get_soup(Url):
    return BeautifulSoup(requests.get(Url).text, 'html.parser')

for s_s in range(1,2):
  Link_one = get_soup(teknosa.format(s_s))
  x = Link_one.find_all("div",{"id":"product-item"})
  for s in x:
      Link_two.append(tekno + s.a['href'])
  for link_site in Link_two:
    computer = get_soup(link_site)

    Title = computer.find("div", {"class":"pdp-base"}).h1.text.strip(" \n").split(" ")
    Marka = Title[0]
    Model_adi = " ".join(Title[1:3])
    Fiyat = computer.find("div", {"class":"prd-prc2"}).text.strip(" \n")
    points = pointss[random.randint(0, 10)]

    row += 1
    Page_urun = computer.find("div", {"class":"pdp-acc-body"}).find("div", {"class":"ptf-body"})
    Ozellikler1 = Page_urun.find_all("table")
    for ozellik in Ozellikler1:
        Ozellikler2 = ozellik.find_all("tr")
        Ozellik_adi = Ozellikler2[0].find_all("th")
        Ozellik_aciklamasi = Ozellikler2[1].find_all("td")

        for i in Ozellik_adi:
            Ozellik_adi2.append(i.text)
        for j in Ozellik_aciklamasi:
            Ozellik_aciklamasi2.append(j.text)
        
    for k in range(len(Ozellik_aciklamasi2)):
        if (Ozellik_adi2[k].find("Model Kodu") != -1):
          Model_no = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("İşletim Sistemi") != -1):
          OS = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("İşlemci Nesli") != -1):
          cpuStatus = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("İşlemci") != -1):
          cpuType = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("Ram") != -1):
          ram = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("SSD Kapasitesi") != -1):
          Disk = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("Disk Türü") != -1):
          DiskType = Ozellik_aciklamasi2[k]
        elif (Ozellik_adi2[k].find("Ekran Boyutu") != -1):
          screen = Ozellik_aciklamasi2[k]
    
    c = my_sheet.cell(row = row, column = 1)
    c.value = Marka
    c = my_sheet.cell(row = row, column = 2)
    c.value = Model_adi
    c = my_sheet.cell(row = row, column = 3)
    c.value = Model_no
    c = my_sheet.cell(row = row, column = 4)
    c.value = OS
    c = my_sheet.cell(row = row, column = 5)
    c.value = cpuType
    c = my_sheet.cell(row = row, column = 6)
    c.value = cpuStatus
    c = my_sheet.cell(row = row, column = 7)
    c.value = ram
    c = my_sheet.cell(row = row, column = 8)
    c.value = Disk
    c = my_sheet.cell(row = row, column = 9)
    c.value = DiskType
    c = my_sheet.cell(row = row, column = 10)
    c.value = screen
    c = my_sheet.cell(row = row, column = 11)
    c.value = points
    c = my_sheet.cell(row = row, column = 12)
    c.value = Fiyat
    c = my_sheet.cell(row = row, column = 13)
    c.value = "teknosa"
    c = my_sheet.cell(row = row, column = 14)
    c.value = link_site

my_wb.save("teknosa.xlsx")
