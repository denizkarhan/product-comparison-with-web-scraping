from gettext import find
from os import link
import requests
from bs4 import BeautifulSoup
import pandas as pd
import xlwt
import openpyxl
from xlwt import Workbook

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

col = ["Marka", "Model Adı", "Model No",
       "İşletim Sistemi", "İşlemci Tipi",
       "İslemci Nesli", "Ram", "Disk Boyutu",
       "Disk Türü", "Ekran Boyutu", "Puanı",
       "Fiyat", "Site İsmi", "Site Linki"]
vatan = "https://www.vatanbilgisayar.com/notebook/?page={0}"
V = "https://www.vatanbilgisayar.com"
OS = " "
cpuType = " "
cpuStatus = " "
ram = " "
Disk = " "
DiskType = " "
screen = " "
Marka = " "
row = 1

for i in range(14):
    c1 = my_sheet.cell(row = 1, column = i + 1)
    c1.value = col[i]

def get_soup(Url):
    return BeautifulSoup(requests.get(Url).text, 'html.parser')

for s_s in range(1,3): 
    page = get_soup(vatan.format(s_s)).find_all("div", {"class":"product-list product-list--list-page"})
    for i in page:
        row += 1
        link_site = V + i.a['href']
        page2 = get_soup(V + i.a['href'])
        puan = str(page2.find("div", {"class":"rank-star"}))
        puan = puan[puan.find("width:") + 6:puan.find("%")]
        points = str(int(puan) / 20)
        Full_Title = page2.find("div", {"class":"product-list__content product-detail-big-price"})
        Fiyat = Full_Title.find("div", {"class":"product-list__cost product-list__description"}).span.text.strip(" \n")
        Title = Full_Title.h1.text.strip(" \n").split(" ")
        Marka = Title[0]
        Model_adi = " ".join(Title[1:3])
        Ozellikler = page2.find_all("div", {"class":"product-feature"})
        key = []
        value = []
        for s in Ozellikler:
            ozellik_adlari = s.find_all("tr")
            for i in ozellik_adlari:
                k1 = i.find_all("td")
                key.append(k1[0].text.strip(" \n"))
                value.append(k1[1].text.strip(" \n"))
        for i in range(len(key)):
            if (key[i].find("İşlemci Teknolojisi") != -1):
                cpuType = value[i]
            elif (key[i].find("İşlemci Nesli") != -1):
                cpuStatus = value[i]
            elif (key[i].find("Ram (Sistem Belleği)") != -1):
                if (value[i].find("(") != -1):
                    r = value[i].find("(")
                    ram = value[i][:r].strip(" \n")
                else:
                    ram = value[i].strip(" \n")
            elif (key[i].find("Ekran Boyutu") != -1):
                screen = value[i]
            elif (key[i].find("Disk Kapasitesi") != -1):
                Disk = value[i][:5]
            elif (key[i].find("Disk Türü") != -1):
                k = 0
                if (value[i].find("NVMe") != -1):
                    k = 4
                DiskType = value[i][k:].strip(" \n")
            elif (key[i].find("İşlemci Numarası") != -1):
                cpuStatus = value[i]
            elif (key[i].find("İşletim Sistemi") != -1):
                OS = value[i]
            elif (key[i].find("Üretici Part Numarası") != -1):
                model_no = value[i]
                
        c = my_sheet.cell(row = row, column = 1)
        c.value = Marka
        c = my_sheet.cell(row = row, column = 2)
        c.value = Model_adi
        c = my_sheet.cell(row = row, column = 3)
        c.value = model_no
        c = my_sheet.cell(row = row, column = 4)
        c.value = OS
        c = my_sheet.cell(row = row, column = 5)
        c.value = cpuType
        c = my_sheet.cell(row = row, column = 6)
        c.value = cpuStatus
        c = my_sheet.cell(row = row, column = 7)
        c.value = ram
        c = my_sheet.cell(row = row, column = 8)
        c.value = Disk
        c = my_sheet.cell(row = row, column = 9)
        c.value = DiskType
        c = my_sheet.cell(row = row, column = 10)
        c.value = screen
        c = my_sheet.cell(row = row, column = 11)
        c.value = points
        c = my_sheet.cell(row = row, column = 12)
        c.value = Fiyat
        c = my_sheet.cell(row = row, column = 13)
        c.value = "vatan"
        c = my_sheet.cell(row = row, column = 14)
        c.value = link_site

my_wb.save("vatan.xlsx")
