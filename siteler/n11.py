from ast import Mod
from gettext import find
import requests
from bs4 import BeautifulSoup
import pandas as pd
import xlwt
import openpyxl
from xlwt import Workbook

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

col = ["Marka", "Model Adı", "Model No", "İşletim Sistemi", "İşlemci Tipi", "İslemci Nesli", "Ram", "Disk Boyutu", "Disk Türü", "Ekran Boyutu", "Puanı", "Fiyat", "Site İsmi", "Site Linki"]
n11 = "https://www.n11.com/bilgisayar/dizustu-bilgisayar?ipg={0}"
OS = " "
cpuType = " "
cpuStatus = " "
ram = " "
Disk = " "
DiskType = " "
screen = " "
Marka = " "
row = 1

for i in range(14):
    c1 = my_sheet.cell(row = 1, column = i + 1)
    c1.value = col[i]

def get_soup(Url):
    return BeautifulSoup(requests.get(Url).text, 'html.parser')

for s_s in range(1,20):
  Link_one = get_soup(n11.format(s_s)).find_all("div", {"class":"pro"})
  for i in Link_one:
    row += 1
    link_site = i.a['href']
    Page_urun = get_soup(i.a['href'])
    ozellikler = Page_urun.find_all("li", {"class":"unf-prop-list-item"})
    fiyat = Page_urun.find("div", {"class":"unf-p-summary-price"}).text
    puan = Page_urun.find("div", {"class":"proRatingHolder"}).find("div", {"class":"ratingCont"}).strong.text
    Title = Page_urun.find("div", {"class":"nameHolder"}).find("h1").text.strip(" \n").split(" ")
    Marka = Title[0]
    Model_adi = Title[3]
    for i in ozellikler:
        str = i.text
        if (i.find("p", {"class":"unf-prop-list-title"}).text.find("Model") != -1 and len(i.find("p", {"class":"unf-prop-list-title"}).text) <= 6):
          ss = i.text[6:].strip(" \n").split(" ")
          if (len(ss) > 1):
            Model_adi = " ".join(ss[:len(ss) - 1]).strip(" \n")
          Model_no = (ss[len(ss) - 1]).strip(" \n")
        elif (str.find("İşletim Sistemi") != -1):
          OS = str[18:].strip(" \n")
        elif (str.find("İşlemci Modeli") != -1):
          cpuStatus = str[17:].strip(" \n")
        elif (str.find("İşlemci") != -1):
          cpuType = str[10:].strip(" \n")
        elif (str.find("Bellek Kapasitesi") != -1 and len(str) < 27):
          ram = str[20:].strip(" \n")
        elif (str.find("Disk Kapasitesi") != -1):
          Disk = str[18:].strip(" \n")
        elif (str.find("Disk Türü") != -1):
          DiskType = str[12:].strip(" \n")
        elif (str.find("Ekran Boyutu") != -1):
          screen = str[15:].strip(" \n")

    c = my_sheet.cell(row = row, column = 1)
    c.value = Marka
    c = my_sheet.cell(row = row, column = 2)
    c.value = Model_adi
    c = my_sheet.cell(row = row, column = 3)
    c.value = Model_no
    c = my_sheet.cell(row = row, column = 4)
    c.value = OS
    c = my_sheet.cell(row = row, column = 5)
    c.value = cpuType
    c = my_sheet.cell(row = row, column = 6)
    c.value = cpuStatus
    c = my_sheet.cell(row = row, column = 7)
    c.value = ram
    c = my_sheet.cell(row = row, column = 8)
    c.value = Disk
    c = my_sheet.cell(row = row, column = 9)
    c.value = DiskType
    c = my_sheet.cell(row = row, column = 10)
    c.value = screen
    c = my_sheet.cell(row = row, column = 11)
    c.value = puan
    c = my_sheet.cell(row = row, column = 12)
    c.value = fiyat
    c = my_sheet.cell(row = row, column = 13)
    c.value = "n11"
    c = my_sheet.cell(row = row, column = 14)
    c.value = link_site

my_wb.save("n11.xlsx")
