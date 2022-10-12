import find
from os import link
import requests
from bs4 import BeautifulSoup
import pandas as pd
import xlwt
import openpyxl
from xlwt import Workbook
import bs4
import lxml

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

col = ["Marka", "Model Adı", "Model No",
       "İşletim Sistemi", "İşlemci Tipi",
       "İslemci Nesli", "Ram", "Disk Boyutu",
       "Disk Türü", "Ekran Boyutu", "Puanı",
       "Fiyat", "Site İsmi", "Site Linki"]
Trendyol = "https://www.trendyol.com/laptop-x-c103108?pi={0}"
OS = " "
cpuType = " "
cpuStatus = " "
ram = " "
Disk = " "
DiskType = " "
screen = " "
row = 1
points_index = 0
full_points = []

def strStr(haystack, needle):
  if needle in haystack:
    return haystack.index(needle)
  return -1

def atoi(str):
    resultant = 0
    for i in range(len(str)):
        if (str[i] < '0' or str[i] > '9'):
          break
        resultant = resultant * 10 + (ord(str[i]) - ord('0'))
    return (resultant)

for i in range(14):
    c1 = my_sheet.cell(row = 1, column = i + 1)
    c1.value = col[i]

def get_soup(Url):
    return BeautifulSoup(requests.get(Url).text, 'html.parser')

for s_s in range(1,2):
  Link_one = get_soup(Trendyol.format(s_s)).find("div", {"class":"prdct-cntnr-wrppr"})
  computers = Link_one.find_all("div", {"class":"p-card-wrppr with-campaign-view"})
  Links_points = Link_one.find_all("div", {"class":"product-down"})
  for s in Links_points:
    rrr = s.find("div", {"class":"ratings"})
    if (str(rrr) == "None"):
      full_points.append(0)
    else:
      pp = rrr.find_all("div", {"class":"star-w"})
      points = 0
      for pp in rrr:
        points += atoi(str(pp)[str(pp).find("style") + 13:])
      full_points.append(points)
  for i in computers:    
    row += 1
    link_site = "https://www.trendyol.com" + i.a['href']
    Page_urun = get_soup("https://www.trendyol.com" + i.a['href'])
    Marka = Page_urun.find("div", {"class":"pr-in-cn"}).h1.a.text

    Model = []
    Model = Page_urun.find("h1", {"class":"pr-new-br"}).span.text.split(" ")
    Model_adi = Model[1] + " " + Model[2]
    
    fiyat = Page_urun.find("span", {"class":"prc-dsc"}).text
    
    c = my_sheet.cell(row = row, column = 1)
    c.value = Marka
    
    c = my_sheet.cell(row = row, column = 2)
    c.value = Model_adi
    
    c = my_sheet.cell(row = row, column = 3)
    if (Marka != "Monster"):
      c.value = Model[-2]
    else:
      c.value = "Default"
  
    ozellikler = Page_urun.find("ul", {"class":"detail-attr-container"}).find_all("li")
    flag = 1
    for i in ozellikler:
      str = i.text
      if (str.find("İşletim Sistemi") != -1):
        OS = str[16:]
      elif (str.find("İşlemci Tipi") != -1):
        cpuType = str[13:]
      elif (str.find("İşlemci Nesli") != -1):
        cpuStatus = str[14:]
      elif (str.find("Ram (Sistem Belleği)") != -1 and len(str) < 27):
        ram = str[20:]
      elif (str.find("SSD") != -1):
        Disk = str[14:]
        DiskType = "SSD"
        flag = 0
      elif (str.find("HDD") != -1 and flag == 1):
        Disk = str[20:]
        DiskType = "HDD"
      elif (str.find("Ekran Boyutu") != -1):
        screen = str[13:]

    c = my_sheet.cell(row = row, column = 4)
    c.value = OS
    c = my_sheet.cell(row = row, column = 5)
    c.value = cpuType
    c = my_sheet.cell(row = row, column = 6)
    c.value = cpuStatus
    c = my_sheet.cell(row = row, column = 7)
    c.value = ram
    c = my_sheet.cell(row = row, column = 8)
    c.value = Disk
    c = my_sheet.cell(row = row, column = 9)
    c.value = DiskType
    c = my_sheet.cell(row = row, column = 10)
    c.value = screen
    c = my_sheet.cell(row = row, column = 11)
    try:
      c.value = full_points[row - 1]/100
    except:
      c.value = 0
    c = my_sheet.cell(row = row, column = 12)
    c.value = fiyat
    c = my_sheet.cell(row = row, column = 13)
    c.value = "Trendyol"
    c = my_sheet.cell(row = row, column = 14)
    c.value = link_site

my_wb.save("Trendyol.xlsx")

# puan = soup.find_all("div", {"class":"full"}, {"style":"width"})
# for z in puan:
#   sayilar.append(int(str(z)[strStr(str(z),"width:") + 6:strStr(str(z),"%")]))
# total = sayilar[kol] + sayilar[kol + 1] + sayilar[kol + 2] + sayilar[kol + 3] + sayilar[kol + 4]
# kol += 5
# total /= 100

""" from database import get_database

dbname = get_database()
collection_name = dbname["user_1_items"]
item_1 = {
  "_id" : "U1IT00001",
  "item_name" : "Blender",
  "max_discount" : "10%",
  "batch_number" : "RR450020FRG",
  "price" : 340,
  "category" : "kitchen appliance"
}

item_2 = {
  "_id" : "U1IT00002",
  "item_name" : "Egg",
  "category" : "food",
  "quantity" : 12,
  "price" : 36,
  "item_description" : "brown country eggs"
}
collection_name.insert_many([item_1,item_2]) """
